from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from fake_useragent import UserAgent
from selenium import webdriver

import os
from time import sleep
from datetime import date
from openpyxl import load_workbook,Workbook

from datetime import date,timedelta
from datetime import datetime

# today_date = date.today()

def driver_instance():
    options = Options()
    options.add_argument("--disable-blink-features")
    options.add_argument("start-maximized")
    #options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--incognito')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    return driver


def storing_data(data_list):

    file_exists =os.path.isfile(os.path.join('results','output-listings.xlsx'))
    if not file_exists :
        create_wb=Workbook()
        sheet=create_wb.active
        sheet.append(['name','title','price','description','date_posted','path','source_url'])
        sheet.append(data_list)
        create_wb.save(os.path.join('results','output-listings.xlsx'))
    else:
        load_wb = load_workbook(os.path.join('results','output-listings.xlsx'))
        sheet = load_wb.active
        sheet.append(data_list)
        load_wb.save(os.path.join('results','output-listings.xlsx'))



def scroll_down_page(driver, number_of_scrolling):

    SCROLL_PAUSE_TIME = 5
    # Get scroll height
    last_height = driver.execute_script("return document.body.scrollHeight")

    for i in range(number_of_scrolling):
        # Scroll down to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # Wait to load page
        sleep(SCROLL_PAUSE_TIME)

        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height



#FORMATING FACEBOOK POSTING TIME TO STANDARD DATE
current_date = date.today()
def formate_date(posted_date):

    if (' at' in posted_date):
        posted_date = posted_date.split(' at')[0]

    if (' h' in posted_date) or (' m' in posted_date):
        return current_date

    elif (' d' in posted_date):
        posted_date = posted_date.split(' d')[0]
        posted_date = current_date - timedelta(days=int(posted_date))
        return posted_date

    elif any(['January' in posted_date, 'February' in posted_date , 'March' in posted_date , 'April' in posted_date , \
        'May' in posted_date ,'June' in posted_date , 'July' in posted_date , 'August'\
        in posted_date , 'September' in posted_date , 'October' in posted_date , 'November' in posted_date , 'December' in posted_date]):
   
        posted_day, month_name = posted_date.split()[0],posted_date.split()[1]
        mnum = datetime.strptime(month_name, '%B').month
        posted_date = str(date.today().year) + '-' + ( "0"+str(mnum) if mnum < 10 else str(mnum) ) +  '-' +( posted_day if int(posted_day) >= 10 else '0'+posted_day)
        return posted_date


